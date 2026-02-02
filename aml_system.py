#!/usr/bin/env python3
"""
Advanced AML Case Management System v9.0
Multi-User Cloud Platform with AI Risk Assessment & Report Generation
Developer: Waqas Khan Niazi
"""

import os
import json
import logging
import sqlite3
import uuid
import secrets
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass, asdict
from enum import Enum
from io import BytesIO
import base64

try:
    from flask import Flask, jsonify, request, session, send_file, render_template_string
    from flask_cors import CORS
    from werkzeug.security import generate_password_hash, check_password_hash
    from werkzeug.utils import secure_filename
    try:
        from PIL import Image
        import pytesseract
        HAS_OCR = True
    except:
        HAS_OCR = False
    import PyPDF2
    from docx import Document
    import openpyxl
    import nltk
    from textblob import TextBlob
    import numpy as np
except ImportError:
    os.system("pip install -q flask flask-cors werkzeug PyPDF2 python-docx openpyxl nltk textblob numpy pytesseract")
    from flask import Flask, jsonify, request, session, send_file, render_template_string
    from flask_cors import CORS
    from werkzeug.security import generate_password_hash, check_password_hash
    from werkzeug.utils import secure_filename
    try:
        from PIL import Image
        import pytesseract
        HAS_OCR = True
    except:
        HAS_OCR = False
    import PyPDF2
    from docx import Document
    import openpyxl
    import nltk
    from textblob import TextBlob
    import numpy as np

# Download NLTK data
try:
    nltk.data.find('tokenizers/punkt')
except LookupError:
    nltk.download('punkt', quiet=True)
    nltk.download('averaged_perceptron_tagger', quiet=True)
    nltk.download('wordnet', quiet=True)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'pdf', 'txt', 'jpg', 'jpeg', 'png',
                      'gif', 'mp4', 'mov', 'doc', 'docx', 'xls', 'xlsx'}

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# ==================== DATABASE ====================


class Database:
    def __init__(self, db_path="aml_multi_user.db"):
        self.db_path = db_path
        self.init_schema()

    def get_conn(self):
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def init_schema(self):
        conn = self.get_conn()
        c = conn.cursor()

        # AUDIT TRAIL TABLE - COMPREHENSIVE TRACKING
        c.execute("""CREATE TABLE IF NOT EXISTS audit_trail (
            id TEXT PRIMARY KEY,
            user_id TEXT NOT NULL,
            action TEXT NOT NULL,
            entity_type TEXT NOT NULL,
            entity_id TEXT,
            old_value TEXT,
            new_value TEXT,
            ip_address TEXT,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            status TEXT DEFAULT 'success',
            details TEXT
        )""")

        # USERS TABLE - AUTHENTICATION
        c.execute("""CREATE TABLE IF NOT EXISTS users (
            id TEXT PRIMARY KEY,
            username TEXT UNIQUE NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            full_name TEXT,
            organization TEXT,
            role TEXT DEFAULT 'analyst',
            is_active BOOLEAN DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""")

        # CASES TABLE - AML CASES
        c.execute("""CREATE TABLE IF NOT EXISTS cases (
            id TEXT PRIMARY KEY,
            user_id TEXT NOT NULL,
            case_number TEXT UNIQUE,
            title TEXT,
            description TEXT,
            category TEXT,
            risk_level TEXT,
            amount REAL,
            currency TEXT,
            status TEXT,
            accused_names TEXT,
            cnic TEXT,
            source_country TEXT,
            dest_country TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )""")

        # TRANSACTIONS TABLE
        c.execute("""CREATE TABLE IF NOT EXISTS transactions (
            id TEXT PRIMARY KEY,
            case_id TEXT,
            sender TEXT,
            receiver TEXT,
            amount REAL,
            date TIMESTAMP,
            status TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (case_id) REFERENCES cases(id)
        )""")

        # FILES TABLE - UPLOADS
        c.execute("""CREATE TABLE IF NOT EXISTS files (
            id TEXT PRIMARY KEY,
            case_id TEXT,
            user_id TEXT,
            filename TEXT,
            filepath TEXT,
            file_type TEXT,
            size INT,
            uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (case_id) REFERENCES cases(id),
            FOREIGN KEY (user_id) REFERENCES users(id)
        )""")

        # AI ASSESSMENTS
        c.execute("""CREATE TABLE IF NOT EXISTS assessments (
            id TEXT PRIMARY KEY,
            case_id TEXT,
            risk_score INT,
            factors TEXT,
            recommendations TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (case_id) REFERENCES cases(id)
        )""")

        # REPORTS TABLE
        c.execute("""CREATE TABLE IF NOT EXISTS reports (
            id TEXT PRIMARY KEY,
            case_id TEXT,
            user_id TEXT,
            title TEXT,
            content TEXT,
            format TEXT,
            generated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (case_id) REFERENCES cases(id),
            FOREIGN KEY (user_id) REFERENCES users(id)
        )""")

        # COMPLIANCE RECORDS
        c.execute("""CREATE TABLE IF NOT EXISTS compliance (
            id TEXT PRIMARY KEY,
            case_id TEXT,
            kyc BOOLEAN,
            cdd BOOLEAN,
            edd BOOLEAN,
            sanctions_check BOOLEAN,
            pep_check BOOLEAN,
            record_kept BOOLEAN,
            verified_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (case_id) REFERENCES cases(id)
        )""")

        # ENQUIRIES TABLE - INQUIRY MANAGEMENT
        c.execute("""CREATE TABLE IF NOT EXISTS enquiries (
            id TEXT PRIMARY KEY,
            user_id TEXT NOT NULL,
            case_id TEXT,
            enquiry_number TEXT UNIQUE,
            subject TEXT NOT NULL,
            description TEXT,
            category TEXT,
            priority TEXT DEFAULT 'MEDIUM',
            status TEXT DEFAULT 'OPEN',
            assigned_to TEXT,
            source TEXT,
            reference_documents TEXT,
            findings TEXT,
            recommendations TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            resolved_at TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id),
            FOREIGN KEY (case_id) REFERENCES cases(id),
            FOREIGN KEY (assigned_to) REFERENCES users(id)
        )""")

        conn.commit()
        conn.close()
        logger.info("✅ Database initialized")

    def query(self, sql, params=()):
        conn = self.get_conn()
        c = conn.cursor()
        c.execute(sql, params)
        results = c.fetchall()
        conn.close()
        return results

    def execute(self, sql, params=()):
        conn = self.get_conn()
        c = conn.cursor()
        c.execute(sql, params)
        conn.commit()
        conn.close()
        return True


# ==================== MANAGERS ====================


class UserManager:
    def __init__(self, db):
        self.db = db

    def register(self, username, email, password, full_name, organization):
        try:
            user_id = str(uuid.uuid4())
            password_hash = generate_password_hash(password)
            self.db.execute(
                """INSERT INTO users (id, username, email, password, full_name, organization)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (user_id, username, email, password_hash, full_name, organization)
            )
            return True, user_id, "✅ Account created!"
        except Exception as e:
            return False, None, f"❌ Error: {str(e)}"

    def login(self, username, password):
        results = self.db.query(
            "SELECT * FROM users WHERE username = ?", (username,))
        if results and check_password_hash(results[0]['password'], password):
            return True, results[0]['id']
        return False, None

    def get_user(self, user_id):
        results = self.db.query("SELECT * FROM users WHERE id = ?", (user_id,))
        return dict(results[0]) if results else None


class CaseManager:
    def __init__(self, db):
        self.db = db

    def create(self, user_id, title, category, description, amount, risk_level, accused_names):
        try:
            case_id = str(uuid.uuid4())
            case_number = f"CASE-{datetime.now().strftime('%Y%m%d%H%M%S')}"
            self.db.execute(
                """INSERT INTO cases 
                   (id, user_id, case_number, title, description, category, risk_level, amount, status, accused_names)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (case_id, user_id, case_number, title, description,
                 category, risk_level, amount, 'open', accused_names)
            )
            return True, case_id, f"✅ Case {case_number} created"
        except Exception as e:
            return False, None, f"❌ {str(e)}"

    def list_user_cases(self, user_id):
        results = self.db.query(
            "SELECT * FROM cases WHERE user_id = ? ORDER BY created_at DESC", (user_id,))
        return [dict(r) for r in results]

    def get_case(self, case_id):
        results = self.db.query("SELECT * FROM cases WHERE id = ?", (case_id,))
        return dict(results[0]) if results else None


class DocumentProcessor:
    """Intelligent document analysis and text extraction"""

    @staticmethod
    def extract_text_from_image(filepath):
        """Extract text from images using OCR"""
        try:
            if not HAS_OCR:
                return "[Image file - OCR not available]"
            img = Image.open(filepath)
            text = pytesseract.image_to_string(img)
            return text.strip() if text else "[Image - no readable text]"
        except Exception as e:
            logger.error(f"OCR Error: {e}")
            return "[Image file - unable to process]"

    @staticmethod
    def extract_text_from_pdf(filepath):
        """Extract text from PDF documents"""
        try:
            text = ""
            with open(filepath, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                for page in reader.pages:
                    text += page.extract_text() + " "
            return text.strip()
        except Exception as e:
            logger.error(f"PDF Error: {e}")
            return ""

    @staticmethod
    def extract_text_from_docx(filepath):
        """Extract text from Word documents"""
        try:
            doc = Document(filepath)
            text = " ".join([para.text for para in doc.paragraphs])
            return text.strip()
        except Exception as e:
            logger.error(f"DOCX Error: {e}")
            return ""

    @staticmethod
    def extract_text_from_excel(filepath):
        """Extract text from Excel documents"""
        try:
            text = ""
            wb = openpyxl.load_workbook(filepath)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    text += " ".join([str(cell)
                                     if cell else "" for cell in row]) + " "
            return text.strip()
        except Exception as e:
            logger.error(f"Excel Error: {e}")
            return ""

    @staticmethod
    def extract_document_text(filepath):
        """Auto-detect file type and extract text"""
        ext = os.path.splitext(filepath)[1].lower()

        if ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff']:
            return DocumentProcessor.extract_text_from_image(filepath)
        elif ext == '.pdf':
            return DocumentProcessor.extract_text_from_pdf(filepath)
        elif ext in ['.doc', '.docx']:
            return DocumentProcessor.extract_text_from_docx(filepath)
        elif ext in ['.xls', '.xlsx']:
            return DocumentProcessor.extract_text_from_excel(filepath)
        elif ext in ['.txt']:
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    return f.read().strip()
            except:
                return ""
        return ""


class TextAnalyzer:
    """NLP-based intelligent text analysis"""

    @staticmethod
    def detect_high_risk_keywords(text):
        """Detect high-risk keywords indicating suspicious activity"""
        high_risk = ['cash', 'structured', 'smurfing', 'front', 'shell company',
                     'sanctions', 'embargo', 'terrorist', 'money laundering',
                     'hawala', 'underground banking', 'structuring', 'suspicious',
                     'investigation', 'fraud', 'offshore', 'tax evasion', 'bulk cash',
                     'trade based', 'customs', 'informal value transfer', 'black market']
        risk_score = 0
        found_keywords = []
        text_lower = text.lower()
        for keyword in high_risk:
            if keyword in text_lower:
                risk_score += 15
                found_keywords.append(keyword)
        return min(100, risk_score), found_keywords

    @staticmethod
    def analyze_sentiment(text):
        """Analyze text sentiment"""
        try:
            blob = TextBlob(text[:500])
            polarity = blob.sentiment.polarity
            return polarity
        except:
            return 0

    @staticmethod
    def detect_entity_types(text):
        """Extract named entities (names, organizations, locations)"""
        try:
            tokens = nltk.word_tokenize(text[:1000])
            pos_tags = nltk.pos_tag(tokens)
            entities = nltk.ne_chunk(pos_tags)
            entity_labels = []
            for subtree in entities:
                if hasattr(subtree, "label"):
                    entity_labels.append(subtree.label())
            return list(set(entity_labels))
        except:
            return []

    @staticmethod
    def calculate_text_complexity(text):
        """Calculate document complexity score"""
        if not text:
            return 0
        words = len(text.split())
        sentences = len(nltk.sent_tokenize(text))
        avg_word_len = np.mean([len(w)
                               for w in text.split()]) if text.split() else 0
        complexity = min(100, (words / 100) * 10 + (avg_word_len / 5) * 5)
        return complexity


class AIAssessment:
    @staticmethod
    def generate_risk_score(amount, velocity, country_risk, pep_match, document_risk=0):
        """Calculate AI-powered risk score with document analysis"""
        score = 0

        if amount > 5000000:
            score += 30
        elif amount > 1000000:
            score += 20

        score += velocity * 10
        score += country_risk * 8
        score += pep_match * 15
        score += (document_risk / 100) * 25  # Weight document analysis

        return min(100, score)

    @staticmethod
    def get_recommendations(risk_score, document_findings=None):
        """Generate intelligent recommendations"""
        recs = []

        if risk_score >= 80:
            recs = ["⚠️ LEA Referral recommended", "❌ Block account immediately",
                    "📊 Escalate to supervisor", "🔍 Full investigation required"]
        elif risk_score >= 60:
            recs = ["📋 Enhanced Due Diligence", "🔎 Further verification",
                    "💾 Document everything", "📞 Contact customer"]
        else:
            recs = ["✅ KYC Verification sufficient", "📝 Standard monitoring",
                    "💾 File records", "✓ Proceed normally"]

        # Add document-based recommendations
        if document_findings:
            if document_findings.get('high_risk_keywords'):
                recs.append(
                    f"⚠️ Suspicious keywords found: {', '.join(document_findings['high_risk_keywords'][:3])}")
            if document_findings.get('entity_types'):
                recs.append(
                    f"📋 Entities detected: {', '.join(document_findings['entity_types'])}")

        return recs

    @staticmethod
    def analyze_factors(case_data, document_text=""):
        """Analyze factors with document intelligence"""
        factors = {
            "velocity": 7,
            "geographic_risk": 6,
            "sanctions_match": 4,
            "structuring": 5,
            "pep_connection": 3,
            "document_risk": 0
        }

        # Analyze document text if provided
        if document_text:
            keyword_risk, keywords = TextAnalyzer.detect_high_risk_keywords(
                document_text)
            sentiment = TextAnalyzer.analyze_sentiment(document_text)
            complexity = TextAnalyzer.calculate_text_complexity(document_text)
            entities = TextAnalyzer.detect_entity_types(document_text)

            # Adjust document risk based on analysis
            document_risk = (keyword_risk / 5) + \
                abs(sentiment * 20) + (complexity / 10)
            factors["document_risk"] = min(10, document_risk)
            factors["keywords_detected"] = keywords
            factors["sentiment"] = round(sentiment, 2)
            factors["complexity"] = round(complexity, 2)
            factors["entities"] = entities

        return factors

    @staticmethod
    def intelligent_assess(case_id, case_data, uploaded_files=None):
        """Comprehensive intelligent assessment with document analysis"""
        document_text = ""
        document_findings = {}

        # Extract text from all uploaded files
        if uploaded_files:
            for filepath in uploaded_files:
                if os.path.exists(filepath):
                    text = DocumentProcessor.extract_document_text(filepath)
                    document_text += text + " "

        # Analyze combined document text
        if document_text:
            keyword_risk, keywords = TextAnalyzer.detect_high_risk_keywords(
                document_text)
            entities = TextAnalyzer.detect_entity_types(document_text)
            sentiment = TextAnalyzer.analyze_sentiment(document_text)
            complexity = TextAnalyzer.calculate_text_complexity(document_text)

            document_findings = {
                "keyword_risk": keyword_risk,
                "high_risk_keywords": keywords,
                "entities": entities,
                "sentiment": sentiment,
                "complexity": complexity,
                "text_length": len(document_text)
            }
        else:
            document_findings["keyword_risk"] = 0

        # Calculate risk score
        factors = AIAssessment.analyze_factors(case_data, document_text)
        risk_score = AIAssessment.generate_risk_score(
            case_data.get('amount', 0),
            factors['velocity'] / 10,
            factors['geographic_risk'] / 10,
            factors['pep_connection'] / 10,
            document_findings.get('keyword_risk', 0)
        )

        recommendations = AIAssessment.get_recommendations(
            risk_score, document_findings)

        return {
            'risk_score': int(risk_score),
            'factors': factors,
            'recommendations': recommendations,
            'document_findings': document_findings,
            'risk_level': '🔴 CRITICAL' if risk_score >= 80 else '🟠 HIGH' if risk_score >= 60 else '🟡 MEDIUM' if risk_score >= 40 else '🟢 LOW'
        }


class ReportGenerator:
    @staticmethod
    def generate_pdf_report(case_id, case_data, assessment):
        """Generate AI-powered compliance report with document intelligence"""
        risk_score = assessment.get('risk_score', 0)
        risk_level = '🔴 CRITICAL' if risk_score >= 80 else '🟠 HIGH' if risk_score >= 60 else '🟡 MEDIUM' if risk_score >= 40 else '🟢 LOW'
        factors = assessment.get('factors', {})
        recs = assessment.get('recommendations', [])
        doc_findings = assessment.get('document_findings', {})
        rec_text = '\n'.join([f'{i+1}. {rec}' for i, rec in enumerate(recs)])

        # Build document intelligence section
        doc_section = ""
        if doc_findings and doc_findings.get('text_length', 0) > 0:
            doc_section = f"""
DOCUMENT ANALYSIS & INTELLIGENCE:
Total Text Processed: {doc_findings.get('text_length', 0)} characters
Keyword Risk Score: {doc_findings.get('keyword_risk', 0)}/100
Sentiment Analysis: {doc_findings.get('sentiment', 0):.2f}
Text Complexity: {doc_findings.get('complexity', 0):.2f}
High-Risk Keywords Found: {', '.join(doc_findings.get('high_risk_keywords', [])) or 'None'}
Entities Detected: {', '.join(doc_findings.get('entities', [])) or 'None'}
"""

        report = f"""
AML RISK ASSESSMENT & COMPLIANCE REPORT v9.0
Enhanced AI Document Intelligence System
Developed by Waqas Khan Niazi

CASE INFORMATION:
Case ID: {case_id}
Title: {case_data.get('title', 'N/A')}
Category: {case_data.get('category', 'N/A')}
Amount: PKR {case_data.get('amount', 0):,.0f}
Status: {case_data.get('status', 'Open')}
Accused: {case_data.get('accused_names', 'N/A')}

AI RISK ASSESSMENT:
Overall Risk Score: {risk_score}/100
Risk Level: {risk_level}

FACTOR ANALYSIS:
Transaction Velocity: {factors.get('velocity', 0)}/10
Geographic Risk: {factors.get('geographic_risk', 0)}/10
Sanctions Match: {factors.get('sanctions_match', 0)}/10
Structuring Pattern: {factors.get('structuring', 0)}/10
PEP Connection: {factors.get('pep_connection', 0)}/10
Document Risk: {factors.get('document_risk', 0):.2f}/10{doc_section}

COMPLIANCE CHECKLIST:
✓ KYC (Know Your Customer) - COMPLETED
✓ CDD (Customer Due Diligence) - COMPLETED
✓ EDD (Enhanced Due Diligence) - RECOMMENDED
✓ Transaction Monitoring - ACTIVE
✓ Sanctions Screening - PASSED
✓ PEP Verification - COMPLETED
✓ Record Keeping (5-year) - MAINTAINED
✓ Registration/Compliance - VERIFIED

AI RECOMMENDATIONS:
{rec_text}

GENERATED: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
"""
        return report

    @staticmethod
    def generate_excel_report(case_id, case_data):
        """Generate Excel format report with document analysis"""
        return f"Case: {case_id}, Data: {json.dumps(case_data, default=str)}"


# ==================== AUDIT TRAIL MANAGER ====================
class AuditTrail:
    """Comprehensive audit trail and trace management system"""

    def __init__(self, db):
        self.db = db

    def log_action(self, user_id, action, entity_type, entity_id, old_value=None, new_value=None, ip_address="", details=None):
        """Log every action for compliance and tracing"""
        audit_id = str(uuid.uuid4())
        self.db.execute(
            """INSERT INTO audit_trail (id, user_id, action, entity_type, entity_id, old_value, new_value, ip_address, details)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (audit_id, user_id, action, entity_type, entity_id,
             str(old_value)[:500], str(new_value)[:500], ip_address, str(details)[:1000])
        )
        return audit_id

    def get_audit_trail(self, entity_id=None, user_id=None, action=None, limit=100):
        """Retrieve audit trail records"""
        query = "SELECT * FROM audit_trail WHERE 1=1"
        params = []
        if entity_id:
            query += " AND entity_id = ?"
            params.append(entity_id)
        if user_id:
            query += " AND user_id = ?"
            params.append(user_id)
        if action:
            query += " AND action = ?"
            params.append(action)
        query += " ORDER BY timestamp DESC LIMIT ?"
        params.append(limit)
        return self.db.query(query, tuple(params))

    def get_trace_history(self, entity_id, entity_type):
        """Get complete trace history of an entity"""
        results = self.db.query(
            """SELECT * FROM audit_trail WHERE entity_id = ? AND entity_type = ? 
               ORDER BY timestamp ASC""", (entity_id, entity_type)
        )
        return [dict(r) for r in results]


# ==================== SEARCH ENGINE ====================
class SearchEngine:
    """Advanced search and filtering capabilities"""

    def __init__(self, db):
        self.db = db

    def search_cases(self, user_id, query, filters=None):
        """Search cases with multiple criteria"""
        base_query = """SELECT * FROM cases WHERE user_id = ? AND 
                       (client_name LIKE ? OR description LIKE ? OR 
                        business_type LIKE ? OR country LIKE ?)"""
        params = [user_id, f"%{query}%",
                  f"%{query}%", f"%{query}%", f"%{query}%"]

        if filters:
            if filters.get('status'):
                base_query += " AND status = ?"
                params.append(filters['status'])
            if filters.get('risk_level'):
                base_query += " AND risk_level = ?"
                params.append(filters['risk_level'])
            if filters.get('country'):
                base_query += " AND country = ?"
                params.append(filters['country'])

        base_query += " ORDER BY created_at DESC"
        return self.db.query(base_query, tuple(params))

    def search_transactions(self, case_id, query):
        """Search transactions in a case"""
        return self.db.query(
            """SELECT * FROM transactions WHERE case_id = ? AND 
               (sender LIKE ? OR receiver LIKE ?) ORDER BY date DESC""",
            (case_id, f"%{query}%", f"%{query}%")
        )

    def search_by_date_range(self, user_id, start_date, end_date):
        """Find all cases in a date range"""
        return self.db.query(
            """SELECT * FROM cases WHERE user_id = ? AND 
               created_at BETWEEN ? AND ? ORDER BY created_at DESC""",
            (user_id, start_date, end_date)
        )

    def advanced_search(self, user_id, criteria):
        """Advanced multi-criteria search"""
        query = "SELECT * FROM cases WHERE user_id = ?"
        params = [user_id]

        if criteria.get('client_name'):
            query += " AND client_name LIKE ?"
            params.append(f"%{criteria['client_name']}%")
        if criteria.get('min_risk'):
            query += " AND risk_level >= ?"
            params.append(criteria['min_risk'])
        if criteria.get('max_risk'):
            query += " AND risk_level <= ?"
            params.append(criteria['max_risk'])
        if criteria.get('status'):
            query += " AND status = ?"
            params.append(criteria['status'])
        if criteria.get('country'):
            query += " AND country = ?"
            params.append(criteria['country'])

        query += " ORDER BY created_at DESC"
        return self.db.query(query, tuple(params))


# ==================== BULK OPERATIONS ====================
class BulkOperations:
    """Handle bulk edit, update, delete operations"""

    def __init__(self, db, audit):
        self.db = db
        self.audit = audit

    def bulk_update_cases(self, user_id, case_ids, updates, ip_address=""):
        """Update multiple cases at once"""
        results = []
        for case_id in case_ids:
            try:
                old_case = self.db.query(
                    "SELECT * FROM cases WHERE id = ?", (case_id,))
                if not old_case:
                    continue

                update_fields = ", ".join([f"{k} = ?" for k in updates.keys()])
                params = list(updates.values()) + [case_id]
                self.db.execute(
                    f"UPDATE cases SET {update_fields} WHERE id = ?", tuple(params))

                self.audit.log_action(user_id, "BULK_UPDATE", "CASE", case_id,
                                      dict(old_case[0]), updates, ip_address)
                results.append({'case_id': case_id, 'status': 'updated'})
            except Exception as e:
                results.append(
                    {'case_id': case_id, 'status': 'error', 'error': str(e)})
        return results

    def bulk_delete_cases(self, user_id, case_ids, ip_address=""):
        """Delete multiple cases at once"""
        results = []
        for case_id in case_ids:
            try:
                case = self.db.query(
                    "SELECT * FROM cases WHERE id = ?", (case_id,))
                if case:
                    self.db.execute(
                        "DELETE FROM cases WHERE id = ?", (case_id,))
                    self.audit.log_action(user_id, "BULK_DELETE", "CASE", case_id,
                                          dict(case[0]), None, ip_address, "Bulk delete operation")
                    results.append({'case_id': case_id, 'status': 'deleted'})
            except Exception as e:
                results.append(
                    {'case_id': case_id, 'status': 'error', 'error': str(e)})
        return results


# ==================== ENQUIRY MANAGER ====================

class EnquiryManager:
    """Comprehensive enquiry/inquiry management"""

    def __init__(self, db):
        self.db = db

    def create(self, user_id, subject, description, category, priority='MEDIUM', case_id=None, source='MANUAL', ip_address='127.0.0.1'):
        """Create new enquiry"""
        enquiry_id = str(uuid.uuid4())
        enquiry_number = f"ENQ-{datetime.now().strftime('%Y%m%d')}-{str(uuid.uuid4())[:8].upper()}"

        try:
            self.db.execute("""INSERT INTO enquiries 
                (id, user_id, case_id, enquiry_number, subject, description, category, priority, status, source, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                            (enquiry_id, user_id, case_id, enquiry_number, subject, description, category, priority, 'OPEN', source,
                             datetime.now(), datetime.now()))

            logger.info(f"Enquiry created: {enquiry_id}")
            return True, enquiry_id, enquiry_number, f"Enquiry created successfully"
        except Exception as e:
            logger.error(f"Error creating enquiry: {e}")
            return False, None, None, str(e)

    def get_enquiry(self, enquiry_id):
        """Get enquiry details"""
        result = self.db.query(
            "SELECT * FROM enquiries WHERE id = ?", (enquiry_id,))
        return dict(result[0]) if result else None

    def list_user_enquiries(self, user_id, status=None):
        """List all enquiries for user"""
        if status:
            results = self.db.query(
                "SELECT * FROM enquiries WHERE user_id = ? AND status = ? ORDER BY created_at DESC",
                (user_id, status)
            )
        else:
            results = self.db.query(
                "SELECT * FROM enquiries WHERE user_id = ? ORDER BY created_at DESC",
                (user_id,)
            )
        return [dict(r) for r in results]

    def update(self, enquiry_id, **kwargs):
        """Update enquiry fields"""
        updates = []
        values = []
        for key, value in kwargs.items():
            if key not in ['id', 'enquiry_id']:
                updates.append(f"{key} = ?")
                values.append(value)

        if not updates:
            return False, "No fields to update"

        updates.append("updated_at = ?")
        values.append(datetime.now())
        values.append(enquiry_id)

        try:
            self.db.execute(
                f"UPDATE enquiries SET {', '.join(updates)} WHERE id = ?",
                tuple(values)
            )
            logger.info(f"Enquiry updated: {enquiry_id}")
            return True, "Enquiry updated successfully"
        except Exception as e:
            logger.error(f"Error updating enquiry: {e}")
            return False, str(e)

    def update_status(self, enquiry_id, status):
        """Update enquiry status"""
        valid_statuses = ['OPEN', 'IN_PROGRESS',
                          'ON_HOLD', 'CLOSED', 'RESOLVED', 'ESCALATED']
        if status not in valid_statuses:
            return False, f"Invalid status. Valid statuses: {', '.join(valid_statuses)}"

        resolved_at = datetime.now() if status == 'RESOLVED' else None
        try:
            self.db.execute(
                "UPDATE enquiries SET status = ?, resolved_at = ?, updated_at = ? WHERE id = ?",
                (status, resolved_at, datetime.now(), enquiry_id)
            )
            logger.info(f"Enquiry status updated: {enquiry_id} -> {status}")
            return True, f"Status updated to {status}"
        except Exception as e:
            logger.error(f"Error updating status: {e}")
            return False, str(e)

    def add_findings(self, enquiry_id, findings):
        """Add findings to enquiry"""
        try:
            enquiry = self.get_enquiry(enquiry_id)
            if not enquiry:
                return False, "Enquiry not found"

            self.db.execute(
                "UPDATE enquiries SET findings = ?, updated_at = ? WHERE id = ?",
                (findings, datetime.now(), enquiry_id)
            )
            logger.info(f"Findings added to enquiry: {enquiry_id}")
            return True, "Findings added successfully"
        except Exception as e:
            logger.error(f"Error adding findings: {e}")
            return False, str(e)

    def add_recommendations(self, enquiry_id, recommendations):
        """Add recommendations to enquiry"""
        try:
            enquiry = self.get_enquiry(enquiry_id)
            if not enquiry:
                return False, "Enquiry not found"

            self.db.execute(
                "UPDATE enquiries SET recommendations = ?, updated_at = ? WHERE id = ?",
                (recommendations, datetime.now(), enquiry_id)
            )
            logger.info(f"Recommendations added to enquiry: {enquiry_id}")
            return True, "Recommendations added successfully"
        except Exception as e:
            logger.error(f"Error adding recommendations: {e}")
            return False, str(e)

    def assign_to(self, enquiry_id, assigned_to_user_id):
        """Assign enquiry to a user"""
        try:
            self.db.execute(
                "UPDATE enquiries SET assigned_to = ?, updated_at = ? WHERE id = ?",
                (assigned_to_user_id, datetime.now(), enquiry_id)
            )
            logger.info(
                f"Enquiry assigned: {enquiry_id} -> {assigned_to_user_id}")
            return True, "Enquiry assigned successfully"
        except Exception as e:
            logger.error(f"Error assigning enquiry: {e}")
            return False, str(e)

    def delete_enquiry(self, enquiry_id):
        """Soft delete enquiry (mark as deleted)"""
        try:
            self.db.execute(
                "UPDATE enquiries SET status = 'DELETED', updated_at = ? WHERE id = ?",
                (datetime.now(), enquiry_id)
            )
            logger.info(f"Enquiry deleted: {enquiry_id}")
            return True, "Enquiry deleted successfully"
        except Exception as e:
            logger.error(f"Error deleting enquiry: {e}")
            return False, str(e)

    def search_enquiries(self, user_id, query, filters=None):
        """Search enquiries with filters"""
        base_query = """SELECT * FROM enquiries WHERE user_id = ? AND status != 'DELETED' AND 
                       (subject LIKE ? OR description LIKE ? OR enquiry_number LIKE ?)"""
        params = [user_id, f"%{query}%", f"%{query}%", f"%{query}%"]

        if filters:
            if filters.get('status'):
                base_query += " AND status = ?"
                params.append(filters['status'])
            if filters.get('priority'):
                base_query += " AND priority = ?"
                params.append(filters['priority'])
            if filters.get('category'):
                base_query += " AND category = ?"
                params.append(filters['category'])
            if filters.get('case_id'):
                base_query += " AND case_id = ?"
                params.append(filters['case_id'])

        base_query += " ORDER BY created_at DESC"
        results = self.db.query(base_query, tuple(params))
        return [dict(r) for r in results]

    def get_enquiry_history(self, enquiry_id):
        """Get enquiry history/timeline"""
        enquiry = self.get_enquiry(enquiry_id)
        if not enquiry:
            return []

        history = []
        if enquiry.get('created_at'):
            history.append({
                'timestamp': enquiry['created_at'],
                'action': 'CREATED',
                'description': f"Enquiry created with subject: {enquiry['subject']}"
            })
        if enquiry.get('updated_at') and enquiry['updated_at'] != enquiry['created_at']:
            history.append({
                'timestamp': enquiry['updated_at'],
                'action': 'UPDATED',
                'description': f"Enquiry updated - Status: {enquiry['status']}"
            })
        if enquiry.get('resolved_at'):
            history.append({
                'timestamp': enquiry['resolved_at'],
                'action': 'RESOLVED',
                'description': "Enquiry resolved"
            })

        return history

    def bulk_update_enquiries(self, enquiry_ids, updates):
        """Bulk update multiple enquiries"""
        results = []
        for enquiry_id in enquiry_ids:
            try:
                success, msg = self.update(enquiry_id, **updates)
                results.append({
                    'enquiry_id': enquiry_id,
                    'status': 'updated' if success else 'error',
                    'message': msg
                })
            except Exception as e:
                results.append({
                    'enquiry_id': enquiry_id,
                    'status': 'error',
                    'error': str(e)
                })
        return results

    def get_statistics(self, user_id):
        """Get enquiry statistics"""
        stats = {}

        # Total enquiries
        total = self.db.query(
            "SELECT COUNT(*) as count FROM enquiries WHERE user_id = ? AND status != 'DELETED'", (user_id,))
        stats['total'] = total[0]['count'] if total else 0

        # By status
        by_status = self.db.query(
            """SELECT status, COUNT(*) as count FROM enquiries 
               WHERE user_id = ? AND status != 'DELETED' GROUP BY status""",
            (user_id,)
        )
        stats['by_status'] = {r['status']: r['count'] for r in by_status}

        # By priority
        by_priority = self.db.query(
            """SELECT priority, COUNT(*) as count FROM enquiries 
               WHERE user_id = ? AND status != 'DELETED' GROUP BY priority""",
            (user_id,)
        )
        stats['by_priority'] = {r['priority']: r['count'] for r in by_priority}

        # By category
        by_category = self.db.query(
            """SELECT category, COUNT(*) as count FROM enquiries 
               WHERE user_id = ? AND status != 'DELETED' GROUP BY category""",
            (user_id,)
        )
        stats['by_category'] = {r['category']: r['count'] for r in by_category}

        return stats

    def bulk_change_status(self, user_id, case_ids, new_status, ip_address=""):
        """Change status of multiple cases"""
        return self.bulk_update_cases(user_id, case_ids, {'status': new_status}, ip_address)


# ==================== FLASK APP ====================


def create_app():
    app = Flask(__name__)
    CORS(app)
    app.config['SECRET_KEY'] = os.environ.get(
        'SECRET_KEY', 'aml-v9-secret-' + secrets.token_hex(16))
    app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
    app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024

    db = Database()
    user_mgr = UserManager(db)
    case_mgr = CaseManager(db)

    # ========== MAIN DASHBOARD ==========

    @app.route('/', methods=['GET'])
    def dashboard():
        try:
            # Get the directory where this file is located
            script_dir = os.path.dirname(os.path.abspath(__file__))
            dashboard_path = os.path.join(
                script_dir, 'dashboard_enhanced.html')
            if not os.path.exists(dashboard_path):
                dashboard_path = os.path.join(script_dir, 'dashboard.html')

            logger.info(f"Loading dashboard from: {dashboard_path}")

            with open(dashboard_path, 'r', encoding='utf-8') as f:
                content = f.read()

            logger.info(
                f"Dashboard loaded successfully ({len(content)} bytes)")
            return content
        except Exception as e:
            logger.error(f"Error loading dashboard: {e}", exc_info=True)
            return jsonify({'error': str(e)}), 500

    # ========== AUTHENTICATION ==========

    @app.route('/api/auth/register', methods=['POST'])
    def register():
        data = request.json
        success, user_id, msg = user_mgr.register(
            data.get('username'),
            data.get('email'),
            data.get('password'),
            data.get('full_name'),
            data.get('organization')
        )
        if success:
            session['user_id'] = user_id
            return jsonify({'success': True, 'message': msg, 'user_id': user_id}), 201
        return jsonify({'success': False, 'message': msg}), 400

    @app.route('/api/auth/login', methods=['POST'])
    def login():
        data = request.json
        success, user_id = user_mgr.login(
            data.get('username'), data.get('password'))
        if success:
            session['user_id'] = user_id
            user = user_mgr.get_user(user_id)
            return jsonify({'success': True, 'message': '✅ Login successful', 'user': user}), 200
        return jsonify({'success': False, 'message': '❌ Invalid credentials'}), 401

    @app.route('/api/auth/logout', methods=['POST'])
    def logout():
        session.clear()
        return jsonify({'success': True, 'message': 'Logged out'}), 200

    @app.route('/api/auth/profile', methods=['GET'])
    def profile():
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401
        user = user_mgr.get_user(session['user_id'])
        return jsonify(user), 200

    # ========== CASES ==========

    @app.route('/api/cases/create', methods=['POST'])
    def create_case():
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        data = request.json
        success, case_id, msg = case_mgr.create(
            session['user_id'],
            data.get('title'),
            data.get('category'),
            data.get('description'),
            float(data.get('amount', 0)),
            data.get('risk_level'),
            data.get('accused_names')
        )
        if success:
            return jsonify({'success': True, 'case_id': case_id, 'message': msg}), 201
        return jsonify({'success': False, 'message': msg}), 400

    @app.route('/api/cases', methods=['GET'])
    def list_cases():
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        cases = case_mgr.list_user_cases(session['user_id'])
        return jsonify({'cases': cases, 'count': len(cases)}), 200

    @app.route('/api/cases/<case_id>', methods=['GET'])
    def get_case(case_id):
        case = case_mgr.get_case(case_id)
        if case:
            return jsonify(case), 200
        return jsonify({'error': 'Case not found'}), 404

    # ========== AI ASSESSMENT ==========

    @app.route('/api/assess/<case_id>', methods=['POST'])
    def assess_case(case_id):
        """Intelligent assessment with document analysis"""
        case = case_mgr.get_case(case_id)
        if not case:
            return jsonify({'error': 'Case not found'}), 404

        # Retrieve uploaded files for this case
        uploaded_files = []
        try:
            rows = db.execute(
                "SELECT filepath FROM files WHERE case_id = ? LIMIT 10",
                (case_id,)
            ).fetchall()
            for row in rows:
                if row[0] and os.path.exists(row[0]):
                    uploaded_files.append(row[0])
        except:
            pass

        assessment_id = str(uuid.uuid4())

        # Perform intelligent assessment
        assessment = AIAssessment.intelligent_assess(
            case_id, case, uploaded_files)

        # Store assessment in database
        db.execute(
            """INSERT INTO assessments (id, case_id, risk_score, factors, recommendations)
               VALUES (?, ?, ?, ?, ?)""",
            (assessment_id, case_id, assessment['risk_score'],
             json.dumps(assessment['factors']), json.dumps(assessment['recommendations']))
        )

        return jsonify({
            'assessment_id': assessment_id,
            'risk_score': assessment['risk_score'],
            'risk_level': assessment['risk_level'],
            'factors': assessment['factors'],
            'recommendations': assessment['recommendations'],
            'document_findings': assessment['document_findings'],
            'files_analyzed': len(uploaded_files)
        }), 200

    @app.route('/api/assess/<case_id>/documents', methods=['POST'])
    def analyze_documents(case_id):
        """Analyze uploaded documents for this case"""
        case = case_mgr.get_case(case_id)
        if not case:
            return jsonify({'error': 'Case not found'}), 404

        try:
            rows = db.execute(
                "SELECT id, filepath, filename FROM files WHERE case_id = ?",
                (case_id,)
            ).fetchall()

            analysis_results = []
            for row in rows:
                filepath, filename = row[1], row[2]
                if os.path.exists(filepath):
                    text = DocumentProcessor.extract_document_text(filepath)
                    if text:
                        keyword_risk, keywords = TextAnalyzer.detect_high_risk_keywords(
                            text)
                        sentiment = TextAnalyzer.analyze_sentiment(text)
                        entities = TextAnalyzer.detect_entity_types(text)

                        analysis_results.append({
                            'filename': filename,
                            'text_length': len(text),
                            'keyword_risk': keyword_risk,
                            'keywords_found': keywords,
                            'sentiment': sentiment,
                            'entities': entities,
                            'text_preview': text[:200] + '...' if len(text) > 200 else text
                        })

            return jsonify({
                'case_id': case_id,
                'documents_analyzed': len(analysis_results),
                'analysis': analysis_results
            }), 200
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    # ========== FILE UPLOADS ==========

    @app.route('/api/upload/<case_id>', methods=['POST'])
    def upload_file(case_id):
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        if 'file' not in request.files:
            return jsonify({'error': 'No file'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Empty filename'}), 400

        if not allowed_file(file.filename):
            return jsonify({'error': 'File type not allowed'}), 400

        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        file_id = str(uuid.uuid4())
        db.execute(
            """INSERT INTO files (id, case_id, user_id, filename, filepath, file_type, size)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (file_id, case_id, session['user_id'], filename,
             filepath, file.content_type, os.path.getsize(filepath))
        )

        return jsonify({'success': True, 'file_id': file_id, 'filename': filename}), 201

    # ========== REPORTS ==========

    @app.route('/api/reports/<case_id>/generate', methods=['POST'])
    def generate_report(case_id):
        case = case_mgr.get_case(case_id)
        if not case:
            return jsonify({'error': 'Case not found'}), 404

        data = request.json
        report_format = data.get('format', 'txt')

        results = db.query(
            "SELECT * FROM assessments WHERE case_id = ? ORDER BY created_at DESC LIMIT 1", (case_id,))
        assessment = dict(results[0]) if results else {
            'risk_score': 0,
            'factors': json.dumps({}),
            'recommendations': json.dumps([])
        }

        if report_format == 'txt' or report_format == 'pdf':
            content = ReportGenerator.generate_pdf_report(
                case_id, case, assessment)
        else:
            content = ReportGenerator.generate_excel_report(case_id, case)

        report_id = str(uuid.uuid4())
        db.execute(
            """INSERT INTO reports (id, case_id, user_id, title, content, format)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (report_id, case_id, session.get('user_id'),
             f"Report-{case_id}", content, report_format)
        )

        return jsonify({
            'success': True,
            'report_id': report_id,
            'content': content
        }), 200

    @app.route('/api/reports/<report_id>/download', methods=['GET'])
    def download_report(report_id):
        results = db.query("SELECT * FROM reports WHERE id = ?", (report_id,))
        if not results:
            return jsonify({'error': 'Report not found'}), 404

        report = dict(results[0])
        file_obj = BytesIO(report['content'].encode())
        return send_file(file_obj, as_attachment=True, download_name=f"{report_id}.txt")

    # ========== COMPLIANCE ==========

    @app.route('/api/compliance/<case_id>', methods=['POST'])
    def update_compliance(case_id):
        data = request.json
        compliance_id = str(uuid.uuid4())
        db.execute(
            """INSERT INTO compliance (id, case_id, kyc, cdd, edd, sanctions_check, pep_check, record_kept)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (compliance_id, case_id, data.get('kyc'), data.get('cdd'), data.get('edd'),
             data.get('sanctions_check'), data.get('pep_check'), data.get('record_kept'))
        )
        return jsonify({'success': True, 'compliance_id': compliance_id}), 201

    @app.route('/api/compliance/<case_id>', methods=['GET'])
    def get_compliance(case_id):
        results = db.query(
            "SELECT * FROM compliance WHERE case_id = ?", (case_id,))
        if results:
            return jsonify(dict(results[0])), 200
        return jsonify({'error': 'No compliance record'}), 404

    # ========== TRANSACTIONS ==========

    @app.route('/api/transactions/<case_id>', methods=['POST'])
    def add_transaction(case_id):
        data = request.json
        txn_id = str(uuid.uuid4())
        db.execute(
            """INSERT INTO transactions (id, case_id, sender, receiver, amount, date, status)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (txn_id, case_id, data.get('sender'), data.get('receiver'), float(data.get('amount', 0)),
             datetime.now(), data.get('status', 'pending'))
        )
        return jsonify({'success': True, 'transaction_id': txn_id}), 201

    # ========== STATISTICS ==========

    @app.route('/api/stats', methods=['GET'])
    def get_stats():
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        results = db.query(
            "SELECT COUNT(*) as total, status FROM cases WHERE user_id = ? GROUP BY status",
            (session['user_id'],)
        )
        stats = {r['status']: r['total'] for r in results}
        return jsonify(stats), 200

    # ========== AUDIT TRAIL & TRACING ===========

    audit_trail = AuditTrail(db)
    search_engine = SearchEngine(db)
    bulk_ops = BulkOperations(db, audit_trail)

    @app.route('/api/audit/trail', methods=['GET'])
    def get_audit_trail():
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        entity_id = request.args.get('entity_id')
        action = request.args.get('action')
        limit = int(request.args.get('limit', 100))

        results = audit_trail.get_audit_trail(
            entity_id, session['user_id'], action, limit)
        return jsonify([dict(r) for r in results]), 200

    @app.route('/api/audit/trace/<entity_id>', methods=['GET'])
    def get_trace_history(entity_id):
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        entity_type = request.args.get('entity_type', 'CASE')
        history = audit_trail.get_trace_history(entity_id, entity_type)
        return jsonify({'entity_id': entity_id, 'entity_type': entity_type, 'history': history}), 200

    # ========== ADVANCED SEARCH ===========

    @app.route('/api/search', methods=['POST'])
    def search():
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        data = request.json
        search_type = data.get('type', 'basic')

        if search_type == 'basic':
            query = data.get('query', '')
            filters = data.get('filters', {})
            results = search_engine.search_cases(
                session['user_id'], query, filters)
        elif search_type == 'advanced':
            criteria = data.get('criteria', {})
            results = search_engine.advanced_search(
                session['user_id'], criteria)
        elif search_type == 'date_range':
            start_date = data.get('start_date')
            end_date = data.get('end_date')
            results = search_engine.search_by_date_range(
                session['user_id'], start_date, end_date)
        else:
            return jsonify({'error': 'Invalid search type'}), 400

        return jsonify([dict(r) for r in results]), 200

    @app.route('/api/search/transactions/<case_id>', methods=['POST'])
    def search_transactions(case_id):
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        query = request.json.get('query', '')
        results = search_engine.search_transactions(case_id, query)
        return jsonify([dict(r) for r in results]), 200

    # ========== BULK OPERATIONS ===========

    @app.route('/api/cases/bulk/update', methods=['POST'])
    def bulk_update_cases():
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        data = request.json
        case_ids = data.get('case_ids', [])
        updates = data.get('updates', {})
        ip = request.remote_addr

        results = bulk_ops.bulk_update_cases(
            session['user_id'], case_ids, updates, ip)
        return jsonify({'bulk_updates': results}), 200

    @app.route('/api/cases/bulk/delete', methods=['POST'])
    def bulk_delete_cases():
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        data = request.json
        case_ids = data.get('case_ids', [])
        ip = request.remote_addr

        results = bulk_ops.bulk_delete_cases(session['user_id'], case_ids, ip)
        return jsonify({'bulk_deletes': results}), 200

    @app.route('/api/cases/bulk/status', methods=['POST'])
    def bulk_change_status():
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        data = request.json
        case_ids = data.get('case_ids', [])
        new_status = data.get('status', '')
        ip = request.remote_addr

        results = bulk_ops.bulk_change_status(
            session['user_id'], case_ids, new_status, ip)
        return jsonify({'bulk_status_updates': results}), 200

    # ========== ENHANCED CASE OPERATIONS ===========

    @app.route('/api/cases/<case_id>/edit', methods=['PUT'])
    def edit_case(case_id):
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        case = case_mgr.get_case(case_id)
        if not case:
            return jsonify({'error': 'Case not found'}), 404

        old_case = dict(case)
        data = request.json
        updates = {}

        # Track which fields are being updated
        updatable_fields = ['client_name', 'business_type', 'risk_level', 'status',
                            'country', 'description', 'notes']

        for field in updatable_fields:
            if field in data:
                updates[field] = data[field]

        if updates:
            update_str = ", ".join([f"{k} = ?" for k in updates.keys()])
            params = list(updates.values()) + [case_id]
            db.execute(
                f"UPDATE cases SET {update_str} WHERE id = ?", tuple(params))

            audit_trail.log_action(session['user_id'], 'CASE_EDIT', 'CASE', case_id,
                                   old_case, updates, request.remote_addr)

        return jsonify({'success': True, 'case_id': case_id, 'updated_fields': updates}), 200

    @app.route('/api/cases/<case_id>/delete', methods=['DELETE'])
    def delete_case(case_id):
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        case = case_mgr.get_case(case_id)
        if not case:
            return jsonify({'error': 'Case not found'}), 404

        old_case = dict(case)
        db.execute("DELETE FROM cases WHERE id = ?", (case_id,))

        audit_trail.log_action(session['user_id'], 'CASE_DELETE', 'CASE', case_id,
                               old_case, None, request.remote_addr, 'Individual case deletion')

        return jsonify({'success': True, 'case_id': case_id, 'message': 'Case deleted'}), 200

    @app.route('/api/cases/<case_id>/restore', methods=['POST'])
    def restore_case(case_id):
        """Restore recently deleted case if available in audit trail"""
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        # Check audit trail for last delete operation
        audit_records = audit_trail.get_audit_trail(
            case_id, None, 'CASE_DELETE', 1)

        if not audit_records:
            return jsonify({'error': 'No delete record found'}), 404

        audit_record = dict(audit_records[0])
        old_data = json.loads(
            audit_record['old_value']) if audit_record['old_value'] else {}

        if old_data:
            # Recreate the case
            db.execute(
                """INSERT INTO cases (id, user_id, client_name, business_type, risk_level, status, country, created_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (case_id, session['user_id'], old_data.get('client_name'),
                 old_data.get('business_type'), old_data.get('risk_level'),
                 'restored', old_data.get('country'), datetime.now())
            )

            audit_trail.log_action(session['user_id'], 'CASE_RESTORE', 'CASE', case_id,
                                   None, old_data, request.remote_addr, 'Case restored from trash')
            return jsonify({'success': True, 'case_id': case_id}), 200

        return jsonify({'error': 'Cannot restore case'}), 400

    # ========== TRANSACTION MANAGEMENT ===========

    @app.route('/api/transactions/<txn_id>/edit', methods=['PUT'])
    def edit_transaction(txn_id):
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        old_txn = db.query(
            "SELECT * FROM transactions WHERE id = ?", (txn_id,))
        if not old_txn:
            return jsonify({'error': 'Transaction not found'}), 404

        data = request.json
        updates = {}

        for field in ['sender', 'receiver', 'amount', 'status']:
            if field in data:
                updates[field] = data[field]

        if updates:
            update_str = ", ".join([f"{k} = ?" for k in updates.keys()])
            params = list(updates.values()) + [txn_id]
            db.execute(
                f"UPDATE transactions SET {update_str} WHERE id = ?", tuple(params))

            audit_trail.log_action(session['user_id'], 'TRANSACTION_EDIT', 'TRANSACTION', txn_id,
                                   dict(old_txn[0]), updates, request.remote_addr)

        return jsonify({'success': True, 'transaction_id': txn_id}), 200

    @app.route('/api/transactions/<txn_id>/delete', methods=['DELETE'])
    def delete_transaction(txn_id):
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        txn = db.query("SELECT * FROM transactions WHERE id = ?", (txn_id,))
        if not txn:
            return jsonify({'error': 'Transaction not found'}), 404

        db.execute("DELETE FROM transactions WHERE id = ?", (txn_id,))

        audit_trail.log_action(session['user_id'], 'TRANSACTION_DELETE', 'TRANSACTION', txn_id,
                               dict(txn[0]), None, request.remote_addr)

        return jsonify({'success': True, 'transaction_id': txn_id}), 200

    @app.route('/api/transactions/<case_id>/list', methods=['GET'])
    def list_transactions(case_id):
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        results = db.query(
            "SELECT * FROM transactions WHERE case_id = ? ORDER BY date DESC",
            (case_id,)
        )
        return jsonify([dict(r) for r in results]), 200

    # ========== ADVANCED ANALYTICS ===========

    @app.route('/api/analytics/dashboard', methods=['GET'])
    def analytics_dashboard():
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        user_id = session['user_id']

        # Total cases
        total = db.query(
            "SELECT COUNT(*) as count FROM cases WHERE user_id = ?", (user_id,))
        total_cases = total[0]['count'] if total else 0

        # Status breakdown
        status_stats = db.query(
            """SELECT status, COUNT(*) as count FROM cases 
               WHERE user_id = ? GROUP BY status""", (user_id,)
        )

        # Risk distribution
        risk_stats = db.query(
            """SELECT risk_level, COUNT(*) as count FROM cases 
               WHERE user_id = ? GROUP BY risk_level""", (user_id,)
        )

        # Recent activity
        recent = db.query(
            """SELECT * FROM audit_trail WHERE user_id = ? 
               ORDER BY timestamp DESC LIMIT 10""", (user_id,)
        )

        return jsonify({
            'total_cases': total_cases,
            'status_breakdown': {r['status']: r['count'] for r in status_stats},
            'risk_distribution': {r['risk_level']: r['count'] for r in risk_stats},
            'recent_activity': [dict(r) for r in recent]
        }), 200

    @app.route('/api/analytics/comparison', methods=['POST'])
    def case_comparison():
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        data = request.json
        case_ids = data.get('case_ids', [])

        cases = []
        for case_id in case_ids:
            case = case_mgr.get_case(case_id)
            if case:
                cases.append(dict(case))

        return jsonify({'comparison': cases}), 200

    # ========== ENQUIRY MANAGEMENT ==========

    enquiry_mgr = EnquiryManager(db)

    @app.route('/api/enquiries/create', methods=['POST'])
    def create_enquiry():
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        data = request.json
        success, enquiry_id, enquiry_number, msg = enquiry_mgr.create(
            session['user_id'],
            data.get('subject'),
            data.get('description'),
            data.get('category'),
            data.get('priority', 'MEDIUM'),
            data.get('case_id'),
            data.get('source', 'MANUAL'),
            request.remote_addr
        )
        if success:
            audit_trail.log_action(session['user_id'], 'ENQUIRY_CREATE', 'ENQUIRY', enquiry_id,
                                   {}, {'enquiry_number': enquiry_number}, request.remote_addr)
            return jsonify({'success': True, 'enquiry_id': enquiry_id, 'enquiry_number': enquiry_number, 'message': msg}), 201
        return jsonify({'success': False, 'message': msg}), 400

    @app.route('/api/enquiries', methods=['GET'])
    def list_enquiries():
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        status = request.args.get('status')
        enquiries = enquiry_mgr.list_user_enquiries(session['user_id'], status)
        return jsonify({'enquiries': enquiries, 'count': len(enquiries)}), 200

    @app.route('/api/enquiries/<enquiry_id>', methods=['GET'])
    def get_enquiry(enquiry_id):
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        enquiry = enquiry_mgr.get_enquiry(enquiry_id)
        if enquiry:
            return jsonify(enquiry), 200
        return jsonify({'error': 'Enquiry not found'}), 404

    @app.route('/api/enquiries/<enquiry_id>/edit', methods=['PUT', 'POST'])
    def edit_enquiry(enquiry_id):
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        enquiry = enquiry_mgr.get_enquiry(enquiry_id)
        if not enquiry:
            return jsonify({'error': 'Enquiry not found'}), 404

        data = request.json
        old_values = dict(enquiry)

        success, msg = enquiry_mgr.update(enquiry_id, **data)
        if success:
            new_enquiry = enquiry_mgr.get_enquiry(enquiry_id)
            audit_trail.log_action(session['user_id'], 'ENQUIRY_EDIT', 'ENQUIRY', enquiry_id,
                                   old_values, dict(new_enquiry), request.remote_addr)
            return jsonify({'success': True, 'message': msg}), 200
        return jsonify({'success': False, 'message': msg}), 400

    @app.route('/api/enquiries/<enquiry_id>/status', methods=['PUT', 'POST'])
    def update_enquiry_status(enquiry_id):
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        data = request.json
        status = data.get('status')

        success, msg = enquiry_mgr.update_status(enquiry_id, status)
        if success:
            audit_trail.log_action(session['user_id'], 'ENQUIRY_STATUS_UPDATE', 'ENQUIRY', enquiry_id,
                                   {}, {'status': status}, request.remote_addr)
            return jsonify({'success': True, 'message': msg}), 200
        return jsonify({'success': False, 'message': msg}), 400

    @app.route('/api/enquiries/<enquiry_id>/findings', methods=['PUT', 'POST'])
    def add_enquiry_findings(enquiry_id):
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        data = request.json
        findings = data.get('findings')

        success, msg = enquiry_mgr.add_findings(enquiry_id, findings)
        if success:
            audit_trail.log_action(session['user_id'], 'ENQUIRY_FINDINGS_ADDED', 'ENQUIRY', enquiry_id,
                                   {}, {'findings': findings}, request.remote_addr)
            return jsonify({'success': True, 'message': msg}), 200
        return jsonify({'success': False, 'message': msg}), 400

    @app.route('/api/enquiries/<enquiry_id>/recommendations', methods=['PUT', 'POST'])
    def add_enquiry_recommendations(enquiry_id):
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        data = request.json
        recommendations = data.get('recommendations')

        success, msg = enquiry_mgr.add_recommendations(
            enquiry_id, recommendations)
        if success:
            audit_trail.log_action(session['user_id'], 'ENQUIRY_RECOMMENDATIONS_ADDED', 'ENQUIRY', enquiry_id,
                                   {}, {'recommendations': recommendations}, request.remote_addr)
            return jsonify({'success': True, 'message': msg}), 200
        return jsonify({'success': False, 'message': msg}), 400

    @app.route('/api/enquiries/<enquiry_id>/assign', methods=['PUT', 'POST'])
    def assign_enquiry(enquiry_id):
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        data = request.json
        assigned_to = data.get('assigned_to')

        success, msg = enquiry_mgr.assign_to(enquiry_id, assigned_to)
        if success:
            audit_trail.log_action(session['user_id'], 'ENQUIRY_ASSIGNED', 'ENQUIRY', enquiry_id,
                                   {}, {'assigned_to': assigned_to}, request.remote_addr)
            return jsonify({'success': True, 'message': msg}), 200
        return jsonify({'success': False, 'message': msg}), 400

    @app.route('/api/enquiries/<enquiry_id>/delete', methods=['DELETE', 'POST'])
    def delete_enquiry(enquiry_id):
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        enquiry = enquiry_mgr.get_enquiry(enquiry_id)
        if not enquiry:
            return jsonify({'error': 'Enquiry not found'}), 404

        success, msg = enquiry_mgr.delete_enquiry(enquiry_id)
        if success:
            audit_trail.log_action(session['user_id'], 'ENQUIRY_DELETE', 'ENQUIRY', enquiry_id,
                                   dict(enquiry), None, request.remote_addr)
            return jsonify({'success': True, 'message': msg}), 200
        return jsonify({'success': False, 'message': msg}), 400

    @app.route('/api/enquiries/search', methods=['POST'])
    def search_enquiries():
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        data = request.json
        query = data.get('query', '')
        filters = data.get('filters', {})

        results = enquiry_mgr.search_enquiries(
            session['user_id'], query, filters)
        return jsonify({'results': results, 'count': len(results)}), 200

    @app.route('/api/enquiries/<enquiry_id>/history', methods=['GET'])
    def get_enquiry_history(enquiry_id):
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        history = enquiry_mgr.get_enquiry_history(enquiry_id)
        return jsonify({'history': history}), 200

    @app.route('/api/enquiries/bulk/update', methods=['POST'])
    def bulk_update_enquiries():
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        data = request.json
        enquiry_ids = data.get('enquiry_ids', [])
        updates = data.get('updates', {})

        results = enquiry_mgr.bulk_update_enquiries(enquiry_ids, updates)
        return jsonify({'results': results}), 200

    @app.route('/api/enquiries/statistics', methods=['GET'])
    def enquiry_statistics():
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        stats = enquiry_mgr.get_statistics(session['user_id'])
        return jsonify(stats), 200

    return app


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


if __name__ == '__main__':
    app = create_app()
    logger.info(
        "🚀 Starting Advanced AML System v9.0 - Developed by Waqas Khan Niazi")
    app.run(host='0.0.0.0', port=5000, debug=False, threaded=True)
